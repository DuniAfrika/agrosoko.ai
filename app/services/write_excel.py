import pandas as pd
from openpyxl import load_workbook
import os
from datetime import datetime

DATA_FILE = "/Users/ratego/Dev/agrosoko.ai/data/prices.xlsx"

def write_to_excel(data):
    """
    Appends scraped data to /data/prices.xlsx.
    Creates a new sheet named YYYY-MM-DD.
    """
    date_str = data.get("date", datetime.now().strftime("%Y-%m-%d"))
    
    df = pd.DataFrame([data])
    
    # Check if file exists
    if not os.path.exists(DATA_FILE):
        # Create new file with the sheet
        with pd.ExcelWriter(DATA_FILE, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=date_str, index=False)
        print(f"Created {DATA_FILE} and added sheet {date_str}")
    else:
        # Append new sheet
        try:
            with pd.ExcelWriter(DATA_FILE, engine='openpyxl', mode='a') as writer:
                if date_str in writer.book.sheetnames:
                    # If sheet exists, maybe append or overwrite? 
                    # PRD says "Adds new sheet with today’s date". 
                    # If it exists, we'll overwrite the sheet by removing it first or just let pandas handle it (pandas usually errors or adds 1)
                    # For safety, let's remove if exists or just append.
                    # Simplest for MVP: If exists, replace it.
                    idx = writer.book.sheetnames.index(date_str)
                    writer.book.remove(writer.book.worksheets[idx])
                    
                df.to_excel(writer, sheet_name=date_str, index=False)
            print(f"Added sheet {date_str} to {DATA_FILE}")
        except Exception as e:
            print(f"Error writing to Excel: {e}")

if __name__ == "__main__":
    # Test
    mock_data = {"date": "2023-10-27", "tomato": 80, "sukuma": 40, "onion": 100, "cabbage": 30}
    write_to_excel(mock_data)
